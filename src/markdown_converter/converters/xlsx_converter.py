from pathlib import Path
from openpyxl import load_workbook
from markdown_converter.config import ConvertOptions
from markdown_converter.converters.base import BaseConverter
from markdown_converter.utils import normalize_text, table_to_markdown


class XlsxConverter(BaseConverter):
    def convert(self, input_path: Path, options: ConvertOptions) -> str:
        wb = load_workbook(filename=str(input_path), data_only=True)
        parts: list[str] = []

        for ws in wb.worksheets:
            parts.append(f"# Sheet: {ws.title}")
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    rows.append(list(row))
            table_md = table_to_markdown(rows)
            if table_md:
                parts.append(table_md)

        return normalize_text("\n\n".join(parts))
